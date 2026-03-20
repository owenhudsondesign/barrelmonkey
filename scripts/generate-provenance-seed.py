#!/usr/bin/env python3
"""
Generate SQL migration to seed batching_runs, bottling_runs, and batching_run_barrels
from WhiskeySystems XLSX exports.

Usage:
    python3 scripts/generate-provenance-seed.py > supabase/migrations/00003_seed_provenance.sql
"""

import openpyxl
from datetime import datetime

# ============================================================
# Spirit type mapping: WS internal → our enum
# ============================================================
SPIRIT_MAP = {
    'Extract Orange': 'extract',
    'Extract Vanilla': 'extract',
    'Vodka Orange': 'vodka',
    'Vodka Vanilla': 'vodka',
    'Vodka Organic': 'vodka',
    'Vodka Non-Organic': 'vodka',
    'Gin Gale Force R2': 'gin',
    'Rum FdC 3yr White': 'rum',
    'Rum FdC 8yr Blend': 'rum',
    'Rum Hurricane Blend': 'rum',
    'brandy pommeau': 'brandy',
    'Whisky Crisp Maris Otter Malt 19': 'whiskey',
    'Whisky Rum Cask 12Yr': 'whiskey',
    'Whisky Notch 15': 'whiskey',
    'Whisky Notch Single Malt': 'whiskey',
    'Whisky Not Scotch 12yr': 'whiskey',
    'Whisky Organic': 'whiskey',
    'Whisky Rum Cask': 'whiskey',
    'Whisky Not Scotch Keg': 'whiskey',
    'Bourbon Nor-Easter': 'bourbon',
    'Bourbon Straight': 'bourbon',
}


def map_spirit(ws_spirit: str) -> str:
    mapped = SPIRIT_MAP.get(ws_spirit)
    if mapped:
        return mapped
    lower = ws_spirit.lower()
    for key in ('whisky', 'whiskey'):
        if key in lower:
            return 'whiskey'
    for key in ('bourbon',):
        if key in lower:
            return 'bourbon'
    for key in ('vodka',):
        if key in lower:
            return 'vodka'
    for key in ('rum',):
        if key in lower:
            return 'rum'
    for key in ('gin',):
        if key in lower:
            return 'gin'
    for key in ('brandy', 'pommeau'):
        if key in lower:
            return 'brandy'
    for key in ('extract',):
        if key in lower:
            return 'extract'
    return 'other'


def esc(val) -> str:
    """Escape a value for SQL."""
    if val is None:
        return 'NULL'
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, datetime):
        return f"'{val.isoformat()}'"
    s = str(val).replace("'", "''")
    return f"'{s}'"


def parse_bottle_size(product_master: str) -> tuple:
    """Extract bottle_size_ml and bottles_per_case from Product Master like '6/750mL 76PF'."""
    if not product_master:
        return None, None
    # Pattern: "6/750mL" or "12/375mL" or "1/50ml"
    import re
    match = re.search(r'(\d+)/(\d+)m[lL]', product_master)
    if match:
        return int(match.group(2)), int(match.group(1))
    return None, None


def main():
    batching_path = '/Users/owenhudson/Downloads/Batching Runs.xlsx'
    bottling_path = '/Users/owenhudson/Downloads/Bottling Runs.xlsx'

    # ============================================================
    # Parse Batching Runs
    # ============================================================
    wb = openpyxl.load_workbook(batching_path, read_only=True)
    batching_rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    batching_runs = []
    for row in batching_rows[1:]:
        ws_id = row[1]
        timestamp = row[3]
        recipe = row[4]
        ws_spirit = row[6] or ''
        first_spirit = row[7] or ''
        wg_made = row[8]
        total_pg = row[9]
        batch_number = str(row[10])
        run_name = row[11]
        lot_name = row[12]
        dest_tank = row[13]
        src_acct = row[14]
        labor = row[15]
        value = row[16]
        user = row[17]

        spirit = map_spirit(ws_spirit)
        product_name = ws_spirit  # Use WS internal spirit type as product name

        batching_runs.append({
            'ws_id': ws_id,
            'batch_number': batch_number,
            'spirit_type': spirit,
            'product_name': product_name,
            'batch_date': timestamp,
            'wine_gal_out': wg_made,
            'proof_gal_out': total_pg,
            'dest_tank': dest_tank,
            'recipe': recipe,
            'lot_name': lot_name,
            'run_name': run_name,
            'notes': f"WS Recipe: {recipe}" if recipe else None,
        })

    # ============================================================
    # Parse Bottling Runs
    # ============================================================
    wb = openpyxl.load_workbook(bottling_path, read_only=True)
    bottling_rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    bottling_runs = []
    for row in bottling_rows[1:]:
        bottling_number = row[1]
        ws_id = row[2]
        product_lot = row[3]
        start_cs = row[4]
        end_cs = row[5]
        lot_name = row[6]
        timestamp = row[7]
        source_tank = row[8]
        ws_spirit = row[9] or ''
        recipe = row[10]
        product_master = row[11] or ''
        ttb_formula = row[12]
        total_cases = row[13] or 0
        loose_btl = row[14] or 0
        starting_pg = row[15]
        pg_bottled = row[16]
        wg_bottled = row[17]
        pg_loss = row[18]
        pct_loss = row[19]
        labor = row[20]
        total_value = row[21]
        price_cs = row[22]
        price_btl = row[23]
        user = row[24]

        spirit = map_spirit(ws_spirit)
        bottle_size_ml, bottles_per_case = parse_bottle_size(product_master)

        # Parse pack format from product master (e.g., "6/750mL")
        pack_format = None
        if product_master:
            import re
            fmt_match = re.search(r'\d+/\d+m[lL]', product_master)
            if fmt_match:
                pack_format = fmt_match.group()

        bottling_runs.append({
            'ws_id': ws_id,
            'bottling_number': int(bottling_number) if bottling_number else None,
            'product_name': product_master.strip() if product_master else ws_spirit,
            'lot_name': str(product_lot) if product_lot else lot_name,
            'spirit_type': spirit,
            'bottle_date': timestamp,
            'source_tank': source_tank,
            'recipe': recipe,
            'ttb_formula': ttb_formula,
            'starting_pg_in_tank': starting_pg,
            'proof_gal_bottled': pg_bottled,
            'wine_gal_bottled': wg_bottled,
            'pack_format': pack_format,
            'bottle_size_ml': bottle_size_ml,
            'bottles_per_case': bottles_per_case,
            'cases_filled': int(total_cases) if total_cases else 0,
            'loose_bottles': int(loose_btl) if loose_btl else 0,
            'pg_loss': pg_loss if pg_loss and pg_loss != 0 else None,
            'pct_loss': pct_loss if pct_loss and pct_loss != 0 else None,
            'labor_hours': labor if labor and labor != 0 else None,
            'total_value': total_value if total_value and total_value != 0 else None,
            'price_per_case': price_cs if price_cs and price_cs != 0 else None,
            'price_per_bottle': price_btl if price_btl and price_btl != 0 else None,
            'start_cs_serial': str(start_cs) if start_cs else None,
            'end_cs_serial': str(end_cs) if end_cs else None,
            'ws_spirit': ws_spirit,
        })

    # ============================================================
    # Collect all tanks needed
    # ============================================================
    all_tanks = set()
    for br in batching_runs:
        if br['dest_tank']:
            all_tanks.add(br['dest_tank'])
    for btl in bottling_runs:
        if btl['source_tank']:
            all_tanks.add(btl['source_tank'])

    # ============================================================
    # Generate SQL
    # ============================================================
    print("-- ============================================================")
    print("-- BARRELMONKEY — PROVENANCE SEED DATA")
    print("-- Generated from WhiskeySystems XLSX exports")
    print(f"-- Batching Runs: {len(batching_runs)} records")
    print(f"-- Bottling Runs: {len(bottling_runs)} records")
    print("-- ============================================================")
    print()
    print("BEGIN;")
    print()

    # -- Admin user for logged_by (use service role or first admin)
    print("-- Use first admin user as logged_by for seeded records")
    print("DO $$ DECLARE v_admin_id UUID;")
    print("BEGIN")
    print("  SELECT id INTO v_admin_id FROM users WHERE role = 'admin' LIMIT 1;")
    print("  IF v_admin_id IS NULL THEN")
    print("    RAISE EXCEPTION 'No admin user found. Import users first.';")
    print("  END IF;")
    print()

    # -- Ensure all referenced tanks exist
    print("  -- Ensure all referenced tanks exist")
    for tank in sorted(all_tanks):
        cat = 'processing'
        if any(c.isdigit() for c in tank[:3]):
            cat = 'misc'  # barrel-ID-like tanks
        print(f"  INSERT INTO tanks (name, tank_category) VALUES ({esc(tank)}, '{cat}')")
        print(f"    ON CONFLICT (name) DO NOTHING;")
    print()

    # -- Insert batching_runs
    print("  -- ============================================================")
    print("  -- BATCHING RUNS")
    print("  -- ============================================================")
    for br in batching_runs:
        print(f"  INSERT INTO batching_runs (")
        print(f"    batch_number, spirit_type, product_name, batch_date,")
        print(f"    wine_gal_out, proof_gal_out, to_tank_id, notes, logged_by")
        print(f"  ) VALUES (")
        print(f"    {esc(br['batch_number'])}, {esc(br['spirit_type'])}::spirit_type, {esc(br['product_name'])},")
        print(f"    {esc(br['batch_date'])},")
        print(f"    {esc(br['wine_gal_out'])}, {esc(br['proof_gal_out'])},")
        print(f"    (SELECT id FROM tanks WHERE name = {esc(br['dest_tank'])} LIMIT 1),")
        notes = br['notes'] or ''
        if br['run_name']:
            notes = f"Run: {br['run_name']}. {notes}".strip()
        if br['lot_name']:
            notes = f"Lot: {br['lot_name']}. {notes}".strip()
        print(f"    {esc(notes if notes else None)},")
        print(f"    v_admin_id")
        print(f"  ) ON CONFLICT (batch_number) DO NOTHING;")
        print()

    # -- Insert bottling_runs, linking to batching_runs via tank + spirit + date
    print("  -- ============================================================")
    print("  -- BOTTLING RUNS")
    print("  -- ============================================================")
    for btl in bottling_runs:
        print(f"  INSERT INTO bottling_runs (")
        print(f"    bottling_number, product_name, lot_name, spirit_type, bottle_date,")
        print(f"    source_tank_id, recipe, ttb_formula,")
        print(f"    starting_pg_in_tank, proof_gal_bottled, wine_gal_bottled,")
        print(f"    pack_format, bottle_size_ml, bottles_per_case,")
        print(f"    cases_filled, loose_bottles,")
        print(f"    pg_loss, pct_loss, labor_hours,")
        print(f"    total_value, price_per_case, price_per_bottle,")
        print(f"    start_cs_serial, end_cs_serial,")
        print(f"    batching_run_id, logged_by")
        print(f"  ) VALUES (")
        print(f"    {esc(btl['bottling_number'])}, {esc(btl['product_name'])}, {esc(btl['lot_name'])},")
        print(f"    {esc(btl['spirit_type'])}::spirit_type, {esc(btl['bottle_date'])},")
        print(f"    (SELECT id FROM tanks WHERE name = {esc(btl['source_tank'])} LIMIT 1),")
        print(f"    {esc(btl['recipe'])}, {esc(btl['ttb_formula'])},")
        print(f"    {esc(btl['starting_pg_in_tank'])}, {esc(btl['proof_gal_bottled'])}, {esc(btl['wine_gal_bottled'])},")
        print(f"    {esc(btl['pack_format'])}, {esc(btl['bottle_size_ml'])}, {esc(btl['bottles_per_case'])},")
        print(f"    {esc(btl['cases_filled'])}, {esc(btl['loose_bottles'])},")
        print(f"    {esc(btl['pg_loss'])}, {esc(btl['pct_loss'])}, {esc(btl['labor_hours'])},")
        print(f"    {esc(btl['total_value'])}, {esc(btl['price_per_case'])}, {esc(btl['price_per_bottle'])},")
        print(f"    {esc(btl['start_cs_serial'])}, {esc(btl['end_cs_serial'])},")
        # Link to batching run: find the most recent batching run that output to this tank
        # with matching spirit type, before the bottling date
        print(f"    (SELECT br.id FROM batching_runs br")
        print(f"     JOIN tanks t ON t.id = br.to_tank_id")
        print(f"     WHERE t.name = {esc(btl['source_tank'])}")
        print(f"       AND br.batch_date < {esc(btl['bottle_date'])}")
        print(f"     ORDER BY br.batch_date DESC LIMIT 1),")
        print(f"    v_admin_id")
        print(f"  );")
        print()

    # -- Link barrels to batching runs via dump events
    print("  -- ============================================================")
    print("  -- BARREL → BATCH LINKAGE")
    print("  -- Match dumped barrels to batching runs via dump events")
    print("  -- Heuristic: barrel dumped to tank T before batch date,")
    print("  -- batch outputs to tank T, matching spirit type")
    print("  -- ============================================================")
    print()
    print("  INSERT INTO batching_run_barrels (batching_run_id, barrel_id, proof_gal_contributed, wine_gal_contributed)")
    print("  SELECT DISTINCT")
    print("    br.id AS batching_run_id,")
    print("    b.id AS barrel_id,")
    print("    be.proof_gal AS proof_gal_contributed,")
    print("    be.wine_gal AS wine_gal_contributed")
    print("  FROM batching_runs br")
    print("  JOIN barrel_events be ON be.event_type = 'dump'")
    print("  JOIN barrels b ON b.id = be.barrel_id")
    print("  -- Match: barrel dumped to same tank that batching run output to")
    print("  -- (at 888, barrels are dumped into a tank, then that tank is 'batched')")
    print("  WHERE be.to_tank_id = br.to_tank_id")
    print("    AND be.event_date < br.batch_date")
    print("    -- Only link dumps within 90 days before the batch")
    print("    AND be.event_date > br.batch_date - INTERVAL '90 days'")
    print("  -- Avoid duplicates from re-runs")
    print("  ON CONFLICT DO NOTHING;")
    print()

    # Also try linking via dump_batch blend_lot_name matching batching_run lot
    print("  -- Also link via dump_batch lot names where available")
    print("  INSERT INTO batching_run_barrels (batching_run_id, barrel_id, proof_gal_contributed, wine_gal_contributed)")
    print("  SELECT DISTINCT")
    print("    br.id,")
    print("    be.barrel_id,")
    print("    be.proof_gal,")
    print("    be.wine_gal")
    print("  FROM batching_runs br")
    print("  JOIN dump_batches db ON db.blend_lot_name IS NOT NULL")
    print("    AND br.notes LIKE '%' || db.blend_lot_name || '%'")
    print("  JOIN barrel_events be ON be.dump_batch_id = db.id")
    print("  WHERE NOT EXISTS (")
    print("    SELECT 1 FROM batching_run_barrels brb")
    print("    WHERE brb.batching_run_id = br.id AND brb.barrel_id = be.barrel_id")
    print("  )")
    print("  ON CONFLICT DO NOTHING;")
    print()

    print("END $$;")
    print()
    print("-- Summary counts")
    print("DO $$ BEGIN")
    print("  RAISE NOTICE 'Batching runs: %', (SELECT count(*) FROM batching_runs);")
    print("  RAISE NOTICE 'Bottling runs: %', (SELECT count(*) FROM bottling_runs);")
    print("  RAISE NOTICE 'Barrel-batch links: %', (SELECT count(*) FROM batching_run_barrels);")
    print("END $$;")
    print()
    print("COMMIT;")


if __name__ == '__main__':
    main()
